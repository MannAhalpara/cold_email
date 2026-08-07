import os
import uuid
import io
import json
import imaplib
import secrets
import hashlib
import time as time_module
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Optional, List
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Depends, Request, Response
from fastapi.responses import HTMLResponse, Response, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import psycopg2
from psycopg2.extras import RealDictCursor
from cryptography.fernet import Fernet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── System Paths & Modules ───────────────────────────────────────────────────
import sys
BACKEND_DIR = Path(__file__).resolve().parent
ROOT_DIR = BACKEND_DIR.parent
FRONTEND_DIR = ROOT_DIR / "frontend"

if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

# Try loading email scraper/AI modules and send_email helper
try:
    import tinyfish_scraper
    import openrouter as openrouter_module
    SCRAPER_AVAILABLE = True
except ImportError:
    SCRAPER_AVAILABLE = False
    print("[WARNING] tinyfish_scraper/openrouter not available.")

try:
    from send_email import send_otp_email
except ImportError:
    print("[WARNING] send_otp_email function not found in send_email.py.")

# ─── Environment ──────────────────────────────────────────────────────────────
load_dotenv(dotenv_path=ROOT_DIR / ".env", override=True)
if not os.getenv("DATABASE_URL"):
    load_dotenv(dotenv_path=BACKEND_DIR / ".env", override=True)

# ─── FastAPI App ───────────────────────────────────────────────────────────────
app = FastAPI(
    title="Cold Email Platform",
    description="Multi-user professor database + personalized email automation",
    version="3.0.0"
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Encryption Helpers ────────────────────────────────────────────────────────
def get_fernet() -> Fernet:
    key = os.getenv("ENCRYPTION_KEY", "").strip()
    if not key:
        raise RuntimeError("ENCRYPTION_KEY not set in .env")
    return Fernet(key.encode())


def encrypt_text(text: str) -> str:
    return get_fernet().encrypt(text.encode()).decode()


def decrypt_text(encrypted: str) -> str:
    return get_fernet().decrypt(encrypted.encode()).decode()


# ─── DB Connection ─────────────────────────────────────────────────────────────
def get_db_connection():
    url = os.getenv("DATABASE_URL", "").strip()
    if not url:
        raise HTTPException(status_code=500, detail="DATABASE_URL not set in .env")
    try:
        return psycopg2.connect(url, cursor_factory=RealDictCursor)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"DB connection failed: {e}")


# ─── Daily Cleanup ─────────────────────────────────────────────────────────────
def cleanup_expired_auth():
    url = os.getenv("DATABASE_URL", "").strip()
    if not url:
        return
    try:
        conn = psycopg2.connect(url)
        with conn.cursor() as cur:
            cur.execute("DELETE FROM otp_codes WHERE expires_at < CURRENT_TIMESTAMP;")
            cur.execute("DELETE FROM user_sessions WHERE expires_at < CURRENT_TIMESTAMP;")
            conn.commit()
        conn.close()
    except Exception as e:
        print(f"[WARNING] Auth cleanup failed: {e}")


# ─── DB Init & Safe Migration (startup) ────────────────────────────────────────
@app.on_event("startup")
def on_startup():
    url = os.getenv("DATABASE_URL", "").strip()
    if not url:
        print("[WARNING] DATABASE_URL not set. Skipping DB init.")
        return
    try:
        conn = psycopg2.connect(url)
        with conn.cursor() as cur:
            # Users table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id UUID PRIMARY KEY,
                    email TEXT UNIQUE NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                );
            """)

            # OTP codes table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS otp_codes (
                    id UUID PRIMARY KEY,
                    email TEXT NOT NULL,
                    otp_hash TEXT NOT NULL,
                    expires_at TIMESTAMPTZ NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                );
            """)

            # User sessions table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS user_sessions (
                    id UUID PRIMARY KEY,
                    user_id UUID NOT NULL REFERENCES users(id) ON DELETE CASCADE,
                    token_hash TEXT NOT NULL,
                    expires_at TIMESTAMPTZ NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
                );
            """)

            # Professors table (with user_id foreign key)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS professors (
                    id TEXT PRIMARY KEY,
                    professor_name TEXT NOT NULL,
                    email TEXT NOT NULL,
                    university TEXT,
                    department TEXT,
                    designation TEXT,
                    country TEXT,
                    research_area TEXT,
                    lab_name TEXT,
                    lab_link TEXT,
                    webpage_1 TEXT,
                    webpage_2 TEXT,
                    webpage_3 TEXT,
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    user_id UUID REFERENCES users(id) ON DELETE CASCADE
                );
                ALTER TABLE professors ADD COLUMN IF NOT EXISTS user_id UUID REFERENCES users(id) ON DELETE CASCADE;
            """)

            # Email settings (encrypted app password)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS email_settings (
                    id SERIAL PRIMARY KEY,
                    email_address TEXT NOT NULL,
                    app_password_encrypted TEXT NOT NULL,
                    updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    user_id UUID REFERENCES users(id) ON DELETE CASCADE
                );
                ALTER TABLE email_settings ADD COLUMN IF NOT EXISTS user_id UUID REFERENCES users(id) ON DELETE CASCADE;
            """)

            # Email template
            cur.execute("""
                CREATE TABLE IF NOT EXISTS email_template (
                    id SERIAL PRIMARY KEY,
                    template_text TEXT NOT NULL,
                    updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    user_id UUID REFERENCES users(id) ON DELETE CASCADE
                );
                ALTER TABLE email_template ADD COLUMN IF NOT EXISTS user_id UUID REFERENCES users(id) ON DELETE CASCADE;
            """)

            # File attachments
            cur.execute("""
                CREATE TABLE IF NOT EXISTS email_attachments (
                    id TEXT PRIMARY KEY,
                    filename TEXT NOT NULL,
                    file_data BYTEA NOT NULL,
                    mime_type TEXT DEFAULT 'application/octet-stream',
                    file_size INTEGER DEFAULT 0,
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    user_id UUID REFERENCES users(id) ON DELETE CASCADE
                );
                ALTER TABLE email_attachments ADD COLUMN IF NOT EXISTS user_id UUID REFERENCES users(id) ON DELETE CASCADE;
            """)

            # Email send log
            cur.execute("""
                CREATE TABLE IF NOT EXISTS email_logs (
                    id TEXT PRIMARY KEY,
                    professor_id TEXT,
                    professor_name TEXT,
                    university TEXT,
                    subject TEXT,
                    body TEXT,
                    status TEXT DEFAULT 'draft',
                    created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    user_id UUID REFERENCES users(id) ON DELETE CASCADE
                );
                ALTER TABLE email_logs ADD COLUMN IF NOT EXISTS user_id UUID REFERENCES users(id) ON DELETE CASCADE;
            """)

            # User credentials table
            cur.execute("""
                CREATE TABLE IF NOT EXISTS user_credentials (
                    workspace_id UUID PRIMARY KEY,
                    sender_email TEXT,
                    encrypted_password TEXT,
                    user_name TEXT,
                    user_bio TEXT,
                    email_template TEXT,
                    updated_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP,
                    user_id UUID REFERENCES users(id) ON DELETE CASCADE
                );
                ALTER TABLE user_credentials ADD COLUMN IF NOT EXISTS user_id UUID REFERENCES users(id) ON DELETE CASCADE;
            """)

            conn.commit()
        conn.close()
        cleanup_expired_auth()
        print("[INFO] Multi-user DB tables & schema migration completed successfully.")
    except Exception as e:
        print(f"[WARNING] DB init error: {e}")


# ─── Auth Dependency ──────────────────────────────────────────────────────────
def get_current_user(request: Request) -> dict:
    raw_token = request.cookies.get("auth_token")
    if not raw_token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            raw_token = auth_header[7:].strip()

    if not raw_token:
        raise HTTPException(status_code=401, detail="Authentication required")

    token_hash = hashlib.sha256(raw_token.encode("utf-8")).hexdigest()

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT u.id, u.email
                FROM user_sessions s
                JOIN users u ON s.user_id = u.id
                WHERE s.token_hash = %s AND s.expires_at > CURRENT_TIMESTAMP;
                """,
                (token_hash,)
            )
            user = cur.fetchone()
            if not user:
                raise HTTPException(status_code=401, detail="Invalid or expired session")
            return {"id": str(user["id"]), "email": user["email"]}
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# AUTHENTICATION ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

class SendOTPIn(BaseModel):
    email: str


class VerifyOTPIn(BaseModel):
    email: str
    otp: str


@app.post("/api/auth/send-otp")
def request_otp(payload: SendOTPIn):
    email = payload.email.strip().lower()
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Invalid email address.")

    otp = f"{secrets.randbelow(1000000):06d}"
    otp_hash = hashlib.sha256(otp.encode("utf-8")).hexdigest()
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)
    otp_id = str(uuid.uuid4())

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM otp_codes WHERE email = %s;", (email,))
            cur.execute(
                """
                INSERT INTO otp_codes (id, email, otp_hash, expires_at)
                VALUES (%s, %s, %s, %s);
                """,
                (otp_id, email, otp_hash, expires_at),
            )
            conn.commit()

        send_otp_email(email, otp)
        cleanup_expired_auth()
        return {"message": f"OTP verification code sent to {email}"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to send OTP: {e}")
    finally:
        conn.close()


@app.post("/api/auth/verify-otp")
def verify_otp(payload: VerifyOTPIn, request: Request, response: Response):
    email = payload.email.strip().lower()
    otp = payload.otp.strip()
    if not email or not otp:
        raise HTTPException(status_code=400, detail="Email and OTP are required.")

    otp_hash = hashlib.sha256(otp.encode("utf-8")).hexdigest()

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT id FROM otp_codes
                WHERE email = %s AND otp_hash = %s AND expires_at > CURRENT_TIMESTAMP;
                """,
                (email, otp_hash),
            )
            otp_row = cur.fetchone()
            if not otp_row:
                raise HTTPException(status_code=400, detail="Invalid or expired OTP code.")

            cur.execute("DELETE FROM otp_codes WHERE email = %s;", (email,))

            cur.execute("SELECT id, email FROM users WHERE email = %s;", (email,))
            user_row = cur.fetchone()
            if user_row:
                user_id = str(user_row["id"])
            else:
                user_id = str(uuid.uuid4())
                cur.execute(
                    "INSERT INTO users (id, email) VALUES (%s, %s) RETURNING *;",
                    (user_id, email),
                )

            raw_token = secrets.token_hex(32)
            token_hash = hashlib.sha256(raw_token.encode("utf-8")).hexdigest()
            session_id = str(uuid.uuid4())
            expires_at = datetime.now(timezone.utc) + timedelta(days=60)

            cur.execute(
                """
                INSERT INTO user_sessions (id, user_id, token_hash, expires_at)
                VALUES (%s, %s, %s, %s);
                """,
                (session_id, user_id, token_hash, expires_at),
            )
            conn.commit()

        is_secure = request.url.scheme == "https"
        response.set_cookie(
            key="auth_token",
            value=raw_token,
            httponly=True,
            samesite="lax",
            secure=is_secure,
            max_age=60 * 86400,
            path="/"
        )
        return {"success": True, "user": {"id": user_id, "email": email}}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Verification failed: {e}")
    finally:
        conn.close()


@app.get("/api/auth/me")
def get_current_user_profile(current_user: dict = Depends(get_current_user)):
    return {"authenticated": True, "user": current_user}


@app.post("/api/auth/logout")
def logout(request: Request, response: Response):
    raw_token = request.cookies.get("auth_token")
    if not raw_token:
        auth_header = request.headers.get("Authorization", "")
        if auth_header.startswith("Bearer "):
            raw_token = auth_header[7:].strip()

    if raw_token:
        token_hash = hashlib.sha256(raw_token.encode("utf-8")).hexdigest()
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM user_sessions WHERE token_hash = %s;", (token_hash,))
                conn.commit()
        except Exception:
            pass
        finally:
            conn.close()

    response.delete_cookie(key="auth_token", path="/")
    return {"success": True}


# ═══════════════════════════════════════════════════════════════════════════════
# PROFESSOR ENDPOINTS (Scoped by user_id)
# ═══════════════════════════════════════════════════════════════════════════════

class ProfessorCreate(BaseModel):
    professor_name: str = Field(..., min_length=1)
    email: str = Field(..., min_length=1)
    university: str = Field(..., min_length=1)
    department: str = Field(..., min_length=1)
    country: str = Field(..., min_length=1)
    research_area: str = Field(..., min_length=1)
    webpage_1: str = Field(..., min_length=1)
    designation: Optional[str] = ""
    lab_name: Optional[str] = ""
    lab_link: Optional[str] = ""
    webpage_2: Optional[str] = ""
    webpage_3: Optional[str] = ""


class ProfessorUpdate(BaseModel):
    professor_name: Optional[str] = None
    email: Optional[str] = None
    university: Optional[str] = None
    department: Optional[str] = None
    country: Optional[str] = None
    research_area: Optional[str] = None
    webpage_1: Optional[str] = None
    designation: Optional[str] = None
    lab_name: Optional[str] = None
    lab_link: Optional[str] = None
    webpage_2: Optional[str] = None
    webpage_3: Optional[str] = None


@app.get("/api/health")
def health_check():
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as cnt FROM professors;")
            res = cur.fetchone()
            count = res["cnt"] if res else 0
        conn.close()
        return {"status": "online", "db_connected": True, "professor_count": count}
    except Exception as e:
        return {"status": "degraded", "db_connected": False, "error": str(e)}


@app.get("/api/weekly-stats")
def get_weekly_stats(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        now = datetime.now()
        start_of_week = (now - timedelta(days=now.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        days_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        today_idx = now.weekday()
        days_data = []
        with conn.cursor() as cur:
            for i in range(7):
                d_start = start_of_week + timedelta(days=i)
                d_end = d_start + timedelta(days=1)
                cur.execute(
                    "SELECT COUNT(*) as cnt FROM professors WHERE user_id = %s AND created_at >= %s AND created_at < %s;",
                    (current_user["id"], d_start, d_end),
                )
                res = cur.fetchone()
                days_data.append({
                    "day": days_names[i],
                    "count": res["cnt"] if res else 0,
                    "is_today": (i == today_idx),
                })
        return {"days": days_data, "today_index": today_idx, "today_name": days_names[today_idx]}
    finally:
        conn.close()


@app.get("/api/professors/export")
def export_professors_excel(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, professor_name, email, university, department, designation,
                       country, research_area, lab_name, lab_link, webpage_1, webpage_2, webpage_3, created_at
                FROM professors WHERE user_id = %s ORDER BY professor_name ASC;
            """, (current_user["id"],))
            rows = cur.fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Professors"
        headers = [
            "ID", "Professor Name", "Email", "University", "Department", "Designation",
            "Country", "Research Area", "Lab Name", "Lab Link", "Webpage 1", "Webpage 2", "Webpage 3", "Created At",
        ]
        ws.append(headers)

        hfill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        hfont = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = halign

        thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )
        for r in rows:
            ws.append([
                r["id"], r["professor_name"], r["email"], r["university"],
                r["department"], r.get("designation", ""), r["country"], r["research_area"],
                r.get("lab_name", ""), r.get("lab_link", ""), r.get("webpage_1", ""),
                r.get("webpage_2", ""), r.get("webpage_3", ""), str(r.get("created_at", "")),
            ])
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(vertical="center")
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(ml + 4, 12), 40)

        buf = io.BytesIO()
        wb.save(buf)
        filename = f"professors_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    finally:
        conn.close()


@app.get("/api/professors")
def list_professors(q: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            if q and q.strip():
                s = f"%{q.strip()}%"
                cur.execute(
                    """SELECT * FROM professors
                       WHERE user_id = %s AND (
                          professor_name ILIKE %s OR email ILIKE %s OR university ILIKE %s
                          OR department ILIKE %s OR research_area ILIKE %s OR country ILIKE %s OR designation ILIKE %s
                       )
                       ORDER BY professor_name ASC;""",
                    (current_user["id"], s, s, s, s, s, s, s),
                )
            else:
                cur.execute("SELECT * FROM professors WHERE user_id = %s ORDER BY professor_name ASC;", (current_user["id"],))
            return [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()


@app.get("/api/professors/{prof_id}")
def get_professor(prof_id: str, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM professors WHERE id = %s AND user_id = %s;", (prof_id, current_user["id"]))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Professor not found")
            return dict(row)
    finally:
        conn.close()


@app.post("/api/professors", status_code=201)
def create_professor(prof: ProfessorCreate, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        prof_id = f"prof_{uuid.uuid4().hex[:8]}"
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO professors
                   (id, professor_name, email, university, department, designation,
                    country, research_area, lab_name, lab_link, webpage_1, webpage_2, webpage_3, user_id)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING *;""",
                (
                    prof_id,
                    prof.professor_name.strip(),
                    prof.email.strip(),
                    prof.university.strip(),
                    prof.department.strip(),
                    prof.designation.strip() if prof.designation else "",
                    prof.country.strip(),
                    prof.research_area.strip(),
                    prof.lab_name.strip() if prof.lab_name else "",
                    prof.lab_link.strip() if prof.lab_link else "",
                    prof.webpage_1.strip(),
                    prof.webpage_2.strip() if prof.webpage_2 else "",
                    prof.webpage_3.strip() if prof.webpage_3 else "",
                    current_user["id"]
                ),
            )
            new_row = cur.fetchone()
            conn.commit()
            return dict(new_row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create professor: {e}")
    finally:
        conn.close()


@app.put("/api/professors/{prof_id}")
def update_professor(prof_id: str, prof: ProfessorUpdate, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM professors WHERE id = %s AND user_id = %s;", (prof_id, current_user["id"]))
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="Professor not found")
            updates, values = [], []
            for key, val in prof.dict(exclude_unset=True).items():
                if val is not None:
                    updates.append(f"{key} = %s")
                    values.append(val.strip() if isinstance(val, str) else val)
            if not updates:
                return dict(existing)
            values.extend([prof_id, current_user["id"]])
            cur.execute(
                f"UPDATE professors SET {', '.join(updates)} WHERE id = %s AND user_id = %s RETURNING *;",
                tuple(values),
            )
            updated_row = cur.fetchone()
            conn.commit()
            return dict(updated_row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update: {e}")
    finally:
        conn.close()


@app.delete("/api/professors/{prof_id}")
def delete_professor(prof_id: str, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM professors WHERE id = %s AND user_id = %s;", (prof_id, current_user["id"]))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Professor not found")
            cur.execute("DELETE FROM professors WHERE id = %s AND user_id = %s;", (prof_id, current_user["id"]))
            conn.commit()
            return {"message": f"Professor '{prof_id}' deleted successfully."}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete: {e}")
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# EMAIL SETTINGS (Scoped by user_id)
# ═══════════════════════════════════════════════════════════════════════════════

class EmailSettingsIn(BaseModel):
    email_address: str
    app_password: str


@app.get("/api/email-settings")
def get_email_settings(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, email_address, updated_at FROM email_settings WHERE user_id = %s ORDER BY id DESC LIMIT 1;",
                (current_user["id"],)
            )
            row = cur.fetchone()
            if not row:
                return {"configured": False}
            return {
                "configured": True,
                "email_address": row["email_address"],
                "updated_at": str(row["updated_at"]),
            }
    finally:
        conn.close()


@app.post("/api/email-settings")
def save_email_settings(settings: EmailSettingsIn, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        encrypted = encrypt_text(settings.app_password)
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM email_settings WHERE user_id = %s LIMIT 1;", (current_user["id"],))
            existing = cur.fetchone()
            if existing:
                cur.execute(
                    "UPDATE email_settings SET email_address=%s, app_password_encrypted=%s, updated_at=CURRENT_TIMESTAMP WHERE id=%s AND user_id=%s;",
                    (settings.email_address, encrypted, existing["id"], current_user["id"]),
                )
            else:
                cur.execute(
                    "INSERT INTO email_settings (email_address, app_password_encrypted, user_id) VALUES (%s, %s, %s);",
                    (settings.email_address, encrypted, current_user["id"]),
                )
            conn.commit()
        return {"success": True, "email_address": settings.email_address}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# EMAIL TEMPLATE (Scoped by user_id)
# ═══════════════════════════════════════════════════════════════════════════════

DEFAULT_TEMPLATE = (
    "Dear Professor {last_name},\n\n"
    "I hope this message finds you well.\n\n"
    "I am writing to express my strong interest in joining your research group at {university} "
    "for potential research opportunities or graduate studies.\n\n"
    "{personalized_paragraph}\n\n"
    "I have been actively building my technical and research background in Artificial Intelligence, "
    "Machine Learning, and Software Development. Key highlights of my background include:\n\n"
    "    \u2022 Technical Projects \u2013 Experience designing and implementing machine learning models "
    "and intelligent systems.\n\n"
    "    \u2022 Analytical & Research Skills \u2013 Strong foundation in quantitative problem solving "
    "and data analysis.\n\n"
    "I am eager to contribute to your ongoing research and would be deeply grateful for an opportunity "
    "to discuss potential research internship or collaboration roles in your lab.\n\n"
    "I have attached my resume for your reference.\n\n"
    "Thank you very much for your time and consideration.\n\n"
    "Best regards,\n"
    "[Your Name]\n"
    "[Your Degree / Department]\n"
    "[Your Contact Email | Phone]"
)


class EmailTemplateIn(BaseModel):
    template_text: str


@app.get("/api/email-template")
def get_email_template(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT template_text FROM email_template WHERE user_id = %s ORDER BY id DESC LIMIT 1;", (current_user["id"],))
            row = cur.fetchone()
            return {"template_text": row["template_text"] if row else DEFAULT_TEMPLATE}
    finally:
        conn.close()


@app.post("/api/email-template")
def save_email_template(tmpl: EmailTemplateIn, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM email_template WHERE user_id = %s LIMIT 1;", (current_user["id"],))
            existing = cur.fetchone()
            if existing:
                cur.execute(
                    "UPDATE email_template SET template_text=%s, updated_at=CURRENT_TIMESTAMP WHERE id=%s AND user_id=%s;",
                    (tmpl.template_text, existing["id"], current_user["id"]),
                )
            else:
                cur.execute(
                    "INSERT INTO email_template (template_text, user_id) VALUES (%s, %s);",
                    (tmpl.template_text, current_user["id"]),
                )
            conn.commit()
        return {"success": True}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# ATTACHMENTS (Scoped by user_id)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/attachments")
def list_attachments(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, filename, mime_type, file_size, created_at FROM email_attachments WHERE user_id = %s ORDER BY created_at DESC;",
                (current_user["id"],)
            )
            rows = cur.fetchall()
            return [
                {
                    "id": r["id"],
                    "filename": r["filename"],
                    "mime_type": r["mime_type"],
                    "file_size": r["file_size"],
                    "created_at": str(r["created_at"]),
                }
                for r in rows
            ]
    finally:
        conn.close()


@app.post("/api/attachments", status_code=201)
async def upload_attachment(file: UploadFile = File(...), current_user: dict = Depends(get_current_user)):
    content = await file.read()
    att_id = f"att_{uuid.uuid4().hex[:8]}"
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "INSERT INTO email_attachments (id, filename, file_data, mime_type, file_size, user_id) VALUES (%s,%s,%s,%s,%s,%s);",
                (att_id, file.filename, psycopg2.Binary(content), file.content_type or "application/octet-stream", len(content), current_user["id"]),
            )
            conn.commit()
        return {"id": att_id, "filename": file.filename, "file_size": len(content)}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


@app.delete("/api/attachments/{att_id}")
def delete_attachment(att_id: str, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM email_attachments WHERE id = %s AND user_id = %s;", (att_id, current_user["id"]))
            conn.commit()
        return {"success": True}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# EMAIL STATS & LOGS (Scoped by user_id)
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/email-stats")
def get_email_stats(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as cnt FROM professors WHERE user_id = %s;", (current_user["id"],))
            total_prof = cur.fetchone()["cnt"]
            cur.execute("SELECT COUNT(DISTINCT professor_id) as cnt FROM email_logs WHERE user_id = %s AND professor_id IS NOT NULL;", (current_user["id"],))
            emails_sent = cur.fetchone()["cnt"]
        return {"total_professors": total_prof, "emails_sent": emails_sent}
    finally:
        conn.close()


@app.get("/api/email-logs")
def get_email_logs(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, professor_id, professor_name, university, subject, status, created_at FROM email_logs WHERE user_id = %s ORDER BY created_at DESC;",
                (current_user["id"],)
            )
            return [dict(r) for r in cur.fetchall()]
    finally:
        conn.close()


@app.get("/api/email-weekly-stats")
def get_email_weekly_stats(current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        now = datetime.now()
        start_of_week = (now - timedelta(days=now.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        days_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        today_idx = now.weekday()
        days_data = []
        with conn.cursor() as cur:
            for i in range(7):
                d_start = start_of_week + timedelta(days=i)
                d_end = d_start + timedelta(days=1)
                cur.execute(
                    "SELECT COUNT(*) as cnt FROM email_logs WHERE user_id = %s AND created_at >= %s AND created_at < %s;",
                    (current_user["id"], d_start, d_end),
                )
                res = cur.fetchone()
                days_data.append({
                    "day": days_names[i],
                    "count": res["cnt"] if res else 0,
                    "is_today": (i == today_idx),
                })
        return {"days": days_data, "today_index": today_idx, "today_name": days_names[today_idx]}
    finally:
        conn.close()


@app.get("/api/emailed-professors")
def get_emailed_professor_ids(current_user: dict = Depends(get_current_user)):
    """Returns list of professor_ids that have been emailed for current user."""
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT DISTINCT professor_id FROM email_logs WHERE user_id = %s AND professor_id IS NOT NULL;", (current_user["id"],))
            return [r["professor_id"] for r in cur.fetchall()]
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# EMAIL GENERATION ENGINE (SSE Scoped by user_id)
# ═══════════════════════════════════════════════════════════════════════════════

def evt(data: dict) -> str:
    """Format SSE event data string."""
    return f"data: {json.dumps(data)}\n\n"


def _get_salutation_and_last_name(full_name: str, designation: str = ""):
    """Extract appropriate salutation using FULL NAME (e.g., 'Professor Krupali Donda' or 'Dr. Krupali Donda')."""
    full_name = (full_name or "").strip()
    designation = (designation or "").strip()
    if not full_name:
        return "Professor", "Professor"

    parts = full_name.split()
    has_dr = any(p.lower() in ["dr.", "dr", "doctor"] for p in parts) or "dr" in designation.lower()
    has_prof = any(p.lower() in ["prof.", "prof", "professor"] for p in parts)

    clean_parts = [p for p in parts if p.lower() not in ["dr.", "dr", "doctor", "prof.", "prof", "professor"]]
    clean_full_name = " ".join(clean_parts) if clean_parts else full_name
    last_name = clean_full_name

    if has_dr and not has_prof:
        salutation = f"Dr. {clean_full_name}"
    else:
        salutation = f"Professor {clean_full_name}"

    return salutation, clean_full_name


def _safe_format(template_text: str, **kwargs) -> str:
    """Safe formatting of string template with dictionary fallbacks."""
    salutation = kwargs.get("salutation", "")
    last_name = kwargs.get("last_name", "")

    formatted = template_text

    if "Dear {last_name}" in formatted and salutation:
        formatted = formatted.replace("Dear {last_name}", f"Dear {salutation}")

    mapping = {
        "salutation": salutation,
        "last_name": last_name,
        "professor_name": kwargs.get("professor_name", ""),
        "university": kwargs.get("university", "your institution"),
        "department": kwargs.get("department", ""),
        "personalized_paragraph": kwargs.get("personalized_paragraph", ""),
    }

    for key, val in mapping.items():
        formatted = formatted.replace("{" + key + "}", str(val))

    return formatted


def email_sse_generator(prof_id: str, user_id: str):
    db_url = os.getenv("DATABASE_URL", "").strip()
    if not db_url:
        yield evt({"step": "error", "message": "DATABASE_URL not configured."})
        return

    # Step 0: Fetch professor
    prof = None
    try:
        conn = psycopg2.connect(db_url, cursor_factory=RealDictCursor)
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM professors WHERE id = %s AND user_id = %s;", (prof_id, user_id))
            row = cur.fetchone()
            if row:
                prof = dict(row)
        conn.close()
    except Exception as e:
        yield evt({"step": "error", "message": f"DB query error: {e}"})
        return

    if not prof:
        yield evt({"step": "error", "message": f"Professor '{prof_id}' not found."})
        return

    # Step 1: Web Scraping
    yield evt({"step": "scraping", "message": "Scraping professor webpage data..."})
    combined_web_data = {"pages": [], "scraped_count": 0}

    if SCRAPER_AVAILABLE:
        try:
            combined_web_data = tinyfish_scraper.fetch_and_combine_prof_sites(prof)
            n = len(combined_web_data.get("pages", []))
            yield evt({"step": "scraping_done", "message": f"Scraped {n} webpage(s) successfully."})
        except Exception as e:
            yield evt({"step": "scraping_done", "message": f"Partial scrape completed. ({str(e)[:60]})"})
    else:
        yield evt({"step": "scraping_done", "message": "Using database info (scraper unavailable)."})

    # Step 2: AI Writing
    yield evt({"step": "writing", "message": "Writing personalized email paragraph with AI..."})

    subject = f"Research Inquiry & Opportunities – {prof.get('professor_name', '')} ({prof.get('university', '')})"

    # 2a: Load template
    template = DEFAULT_TEMPLATE
    try:
        conn2 = psycopg2.connect(db_url, cursor_factory=RealDictCursor)
        with conn2.cursor() as cur:
            cur.execute("SELECT template_text FROM email_template WHERE user_id = %s ORDER BY id DESC LIMIT 1;", (user_id,))
            tmpl_row = cur.fetchone()
        conn2.close()
        if tmpl_row and tmpl_row.get("template_text"):
            template = tmpl_row["template_text"]
    except Exception:
        pass

    # 2b: Generate personalized paragraph
    generic_para = (
        f"Your research on {prof.get('research_area', 'this field')} particularly interests me. "
        "My experience building LLM-based systems, retrieval-augmented generation pipelines, and "
        "applied AI assistants aligns closely with the direction of your work, and I am eager "
        "to contribute meaningfully to your research group."
    )
    personalized = generic_para
    ai_note = "Used generic paragraph (AI unavailable)."

    if SCRAPER_AVAILABLE:
        try:
            personalized = openrouter_module.generate_personalized_paragraph(
                prof=prof,
                combined_web_data=combined_web_data,
                resume_text="",
            )
            ai_note = "AI paragraph generated successfully."
        except Exception as e:
            ai_note = f"AI failed — using generic paragraph. ({str(e)[:80]})"

    # 2c: Assemble full email
    salutation, raw_last_name = _get_salutation_and_last_name(
        prof.get("professor_name", ""),
        prof.get("designation", "")
    )
    try:
        body = _safe_format(
            template,
            salutation=salutation,
            last_name=raw_last_name,
            professor_name=prof.get("professor_name", ""),
            university=prof.get("university", "your institution"),
            department=prof.get("department", ""),
            personalized_paragraph=personalized,
        )
    except Exception:
        body = template

    yield evt({"step": "writing_done", "message": ai_note})

    # Step 3: Attachments
    yield evt({"step": "attaching", "message": "Preparing file attachments..."})
    attachments = []
    try:
        conn3 = psycopg2.connect(db_url, cursor_factory=RealDictCursor)
        with conn3.cursor() as cur:
            cur.execute(
                "SELECT id, filename, mime_type, file_size FROM email_attachments WHERE user_id = %s ORDER BY created_at DESC;",
                (user_id,)
            )
            attachments = [
                {"id": r["id"], "filename": r["filename"], "mime_type": r["mime_type"], "file_size": r["file_size"]}
                for r in cur.fetchall()
            ]
        conn3.close()
    except Exception:
        pass

    yield evt({"step": "attaching_done", "message": f"{len(attachments)} attachment(s) ready."})

    # Step 4: Ready
    yield evt({
        "step": "ready",
        "subject": subject,
        "body": body,
        "to_email": prof.get("email", ""),
        "professor_name": prof.get("professor_name", ""),
        "university": prof.get("university", ""),
        "professor_id": prof_id,
        "attachments": attachments,
    })


@app.post("/api/send-email/{prof_id}")
async def send_email_stream(prof_id: str, current_user: dict = Depends(get_current_user)):
    return StreamingResponse(
        email_sse_generator(prof_id, current_user["id"]),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


# ═══════════════════════════════════════════════════════════════════════════════
# SAVE DRAFT TO GMAIL (Scoped by user_id)
# ═══════════════════════════════════════════════════════════════════════════════

class SaveDraftIn(BaseModel):
    professor_id: str
    professor_name: str
    university: str
    to_email: str
    subject: str
    body: str


@app.post("/api/save-draft")
def save_draft(payload: SaveDraftIn, current_user: dict = Depends(get_current_user)):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT email_address, app_password_encrypted FROM email_settings WHERE user_id = %s ORDER BY id DESC LIMIT 1;",
                (current_user["id"],)
            )
            creds = cur.fetchone()
            if not creds:
                raise HTTPException(status_code=400, detail="Email credentials not configured. Please set up in Settings.")

            sender_email = creds["email_address"]
            app_password = decrypt_text(creds["app_password_encrypted"])

            cur.execute(
                "SELECT filename, file_data, mime_type FROM email_attachments WHERE user_id = %s ORDER BY created_at DESC;",
                (current_user["id"],)
            )
            attachments = cur.fetchall()

        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = payload.to_email
        msg["Subject"] = payload.subject
        msg.attach(MIMEText(payload.body, "plain", "utf-8"))

        for att in attachments:
            part = MIMEApplication(bytes(att["file_data"]), Name=att["filename"])
            part["Content-Disposition"] = f'attachment; filename="{att["filename"]}"'
            msg.attach(part)

        now_imap = imaplib.Time2Internaldate(time_module.time())
        with imaplib.IMAP4_SSL("imap.gmail.com", 993) as imap:
            imap.login(sender_email, app_password)
            res, _ = imap.append('"[Gmail]/Drafts"', r"(\Draft)", now_imap, msg.as_bytes())
            if res != "OK":
                res, data = imap.append("Drafts", r"(\Draft)", now_imap, msg.as_bytes())
                if res != "OK":
                    raise RuntimeError(f"IMAP append failed: {data}")

        log_id = f"log_{uuid.uuid4().hex[:8]}"
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO email_logs (id, professor_id, professor_name, university, subject, body, status, user_id)
                   VALUES (%s,%s,%s,%s,%s,%s,'draft',%s);""",
                (log_id, payload.professor_id, payload.professor_name, payload.university, payload.subject, payload.body, current_user["id"]),
            )
            conn.commit()

        return {"success": True, "log_id": log_id}

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        conn.close()


# ═══════════════════════════════════════════════════════════════════════════════
# SERVE FRONTEND
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/{full_path:path}", response_class=HTMLResponse)
def serve_frontend_routes(full_path: str):
    if full_path.startswith("api/"):
        raise HTTPException(status_code=404, detail="API endpoint not found")
    html_file = FRONTEND_DIR / "index.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding="utf-8"))
    return HTMLResponse(content="<h1>index.html not found</h1>", status_code=404)


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    print(f"[INFO] Cold Email Platform backend starting on http://localhost:{port}")
    uvicorn.run("backend.main:app", host="0.0.0.0", port=port, reload=True)
