import os
import uuid
import csv
import io
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict, Any
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import HTMLResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import psycopg2
from psycopg2.extras import RealDictCursor

# Load environment variables from .env in current folder
BASE_DIR = Path(__file__).resolve().parent
env_path = BASE_DIR / ".env"
if env_path.exists():
    load_dotenv(dotenv_path=env_path, override=True)
else:
    load_dotenv()

app = FastAPI(
    title="Professors Database API",
    description="API for managing academic professors database in Neon PostgreSQL",
    version="1.0.0"
)

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def get_db_url() -> str:
    url = os.getenv("DATABASE_URL", "").strip()
    if not url:
        raise HTTPException(
            status_code=500,
            detail="DATABASE_URL environment variable is missing in .env file."
        )
    return url


def get_db_connection():
    url = get_db_url()
    try:
        conn = psycopg2.connect(url, cursor_factory=RealDictCursor)
        return conn
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to connect to Neon PostgreSQL database: {str(e)}"
        )


def init_db_table():
    url = os.getenv("DATABASE_URL", "").strip()
    if not url:
        print("[WARNING] DATABASE_URL is not set. Skipping table initialization.")
        return False
    try:
        conn = psycopg2.connect(url)
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS professors (
                    id TEXT PRIMARY KEY,
                    professor_name TEXT NOT NULL,
                    email TEXT NOT NULL,
                    university TEXT,
                    department TEXT,
                    designation TEXT,
                    country TEXT,
                    research_area TEXT,
                    lab_name TEXT,
                    lab_link TEXT,
                    webpage_1 TEXT,
                    webpage_2 TEXT,
                    webpage_3 TEXT,
                    created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP
                );
                ALTER TABLE professors ADD COLUMN IF NOT EXISTS created_at TIMESTAMP WITH TIME ZONE DEFAULT CURRENT_TIMESTAMP;
            """)
            conn.commit()
        conn.close()
        print("[INFO] Database table 'professors' initialized successfully in Neon DB.")
        return True
    except Exception as e:
        print(f"[WARNING] Error initializing DB: {e}")
        return False


@app.on_event("startup")
def on_startup():
    init_db_table()


# Pydantic Schemas (Updated validation requirements)
class ProfessorCreate(BaseModel):
    professor_name: str = Field(..., min_length=1)
    email: str = Field(..., min_length=1)
    university: str = Field(..., min_length=1)
    department: str = Field(..., min_length=1)
    country: str = Field(..., min_length=1)
    research_area: str = Field(..., min_length=1)
    webpage_1: str = Field(..., min_length=1)
    
    # Optional fields
    designation: Optional[str] = ""
    lab_name: Optional[str] = ""
    lab_link: Optional[str] = ""
    webpage_2: Optional[str] = ""
    webpage_3: Optional[str] = ""


class ProfessorUpdate(BaseModel):
    professor_name: Optional[str] = None
    email: Optional[str] = None
    university: Optional[str] = None
    department: Optional[str] = None
    country: Optional[str] = None
    research_area: Optional[str] = None
    webpage_1: Optional[str] = None
    designation: Optional[str] = None
    lab_name: Optional[str] = None
    lab_link: Optional[str] = None
    webpage_2: Optional[str] = None
    webpage_3: Optional[str] = None


@app.get("/api/health")
def health_check():
    db_connected = False
    count = 0
    error_msg = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) as cnt FROM professors;")
            res = cur.fetchone()
            count = res["cnt"] if res else 0
        conn.close()
        db_connected = True
    except Exception as e:
        error_msg = str(e)

    return {
        "status": "online" if db_connected else "degraded",
        "db_connected": db_connected,
        "professor_count": count,
        "error": error_msg
    }


@app.get("/api/weekly-stats")
def get_weekly_stats():
    conn = get_db_connection()
    try:
        now = datetime.now()
        # Get start of current week (Monday)
        start_of_week = now - timedelta(days=now.weekday())
        start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        
        days_data = []
        days_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        today_idx = now.weekday()

        with conn.cursor() as cur:
            for i in range(7):
                day_start = start_of_week + timedelta(days=i)
                day_end = day_start + timedelta(days=1)
                
                cur.execute("""
                    SELECT COUNT(*) as cnt FROM professors 
                    WHERE created_at >= %s AND created_at < %s;
                """, (day_start, day_end))
                res = cur.fetchone()
                cnt = res["cnt"] if res else 0
                
                days_data.append({
                    "day": days_names[i],
                    "full_date": day_start.strftime("%Y-%m-%d"),
                    "count": cnt,
                    "is_today": (i == today_idx)
                })

        return {
            "days": days_data,
            "today_index": today_idx,
            "today_name": days_names[today_idx]
        }
    finally:
        conn.close()


@app.get("/api/professors/export")
def export_professors_excel():
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, professor_name, email, university, department, designation, 
                       country, research_area, lab_name, lab_link, webpage_1, webpage_2, webpage_3, created_at
                FROM professors ORDER BY professor_name ASC;
            """)
            rows = cur.fetchall()

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Professors"

        headers = [
            "ID", "Professor Name", "Email", "University", "Department", "Designation",
            "Country", "Research Area", "Lab Name", "Lab Link", "Webpage 1", "Webpage 2", "Webpage 3", "Created At"
        ]
        ws.append(headers)

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        for r in rows:
            ws.append([
                r["id"],
                r["professor_name"],
                r["email"],
                r["university"],
                r["department"],
                r["designation"] or "",
                r["country"],
                r["research_area"],
                r["lab_name"] or "",
                r["lab_link"] or "",
                r["webpage_1"] or "",
                r["webpage_2"] or "",
                r["webpage_3"] or "",
                str(r.get("created_at", ""))
            ])

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 40)

        excel_io = io.BytesIO()
        wb.save(excel_io)
        excel_bytes = excel_io.getvalue()

        filename = f"professors_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    finally:
        conn.close()


@app.get("/api/professors")
def list_professors(q: Optional[str] = None):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            if q and q.strip():
                search = f"%{q.strip()}%"
                query = """
                    SELECT * FROM professors
                    WHERE professor_name ILIKE %s
                       OR email ILIKE %s
                       OR university ILIKE %s
                       OR department ILIKE %s
                       OR research_area ILIKE %s
                       OR country ILIKE %s
                       OR designation ILIKE %s
                    ORDER BY professor_name ASC;
                """
                cur.execute(query, (search, search, search, search, search, search, search))
            else:
                cur.execute("SELECT * FROM professors ORDER BY professor_name ASC;")
            
            rows = cur.fetchall()
            return [dict(row) for row in rows]
    finally:
        conn.close()


@app.get("/api/professors/{prof_id}")
def get_professor(prof_id: str):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM professors WHERE id = %s;", (prof_id,))
            row = cur.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="Professor not found")
            return dict(row)
    finally:
        conn.close()


@app.post("/api/professors", status_code=201)
def create_professor(prof: ProfessorCreate):
    conn = get_db_connection()
    try:
        prof_id = f"prof_{uuid.uuid4().hex[:8]}"
        with conn.cursor() as cur:
            query = """
                INSERT INTO professors (
                    id, professor_name, email, university, department, designation,
                    country, research_area, lab_name, lab_link, webpage_1, webpage_2, webpage_3
                ) VALUES (
                    %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                ) RETURNING *;
            """
            cur.execute(query, (
                prof_id,
                prof.professor_name.strip(),
                prof.email.strip(),
                prof.university.strip(),
                prof.department.strip(),
                prof.designation.strip() if prof.designation else "",
                prof.country.strip(),
                prof.research_area.strip(),
                prof.lab_name.strip() if prof.lab_name else "",
                prof.lab_link.strip() if prof.lab_link else "",
                prof.webpage_1.strip(),
                prof.webpage_2.strip() if prof.webpage_2 else "",
                prof.webpage_3.strip() if prof.webpage_3 else ""
            ))
            new_row = cur.fetchone()
            conn.commit()
            return dict(new_row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to create professor: {str(e)}")
    finally:
        conn.close()


@app.put("/api/professors/{prof_id}")
def update_professor(prof_id: str, prof: ProfessorUpdate):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM professors WHERE id = %s;", (prof_id,))
            existing = cur.fetchone()
            if not existing:
                raise HTTPException(status_code=404, detail="Professor not found")

            updates = []
            values = []
            data = prof.dict(exclude_unset=True)

            for key, val in data.items():
                if val is not None:
                    updates.append(f"{key} = %s")
                    values.append(val.strip() if isinstance(val, str) else val)

            if not updates:
                return dict(existing)

            values.append(prof_id)
            query = f"UPDATE professors SET {', '.join(updates)} WHERE id = %s RETURNING *;"
            cur.execute(query, tuple(values))
            updated_row = cur.fetchone()
            conn.commit()
            return dict(updated_row)
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to update professor: {str(e)}")
    finally:
        conn.close()


@app.delete("/api/professors/{prof_id}")
def delete_professor(prof_id: str):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT id FROM professors WHERE id = %s;", (prof_id,))
            if not cur.fetchone():
                raise HTTPException(status_code=404, detail="Professor not found")

            cur.execute("DELETE FROM professors WHERE id = %s;", (prof_id,))
            conn.commit()
            return {"message": f"Professor '{prof_id}' deleted successfully."}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to delete professor: {str(e)}")
    finally:
        conn.close()


@app.get("/", response_class=HTMLResponse)
def serve_index():
    html_file = BASE_DIR / "index.html"
    if html_file.exists():
        return HTMLResponse(content=html_file.read_text(encoding="utf-8"))
    return HTMLResponse(content="<h1>Index.html not found</h1>", status_code=404)


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    print(f"[INFO] Starting Professors DB Server on http://localhost:{port}")
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
